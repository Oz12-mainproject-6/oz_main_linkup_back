import httpx
import pandas as pd
import io
import asyncio
import re
from typing import BinaryIO
from typing import Any
from fastapi import HTTPException, UploadFile
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.features.events.schemas import FileUploadResponse
from datetime import datetime, timedelta
from tortoise.exceptions import DoesNotExist
from app.features.artists.models import Artist
from app.features.events.models import EventCategory, Events, EventVisibility

# author : Juwon
# date : 2025.09.22
# content : crud.py를 services.py로 이동
class EventCRUD:
    """이벤트 CRUD 클래스 (조회, 일괄 생성, 알림 관련만 남김)"""

    @staticmethod
    async def get_list(
        skip: int = 0,
        limit: int = 100,
        artist_parent_group: int | None = None,  # 🔹 Artist의 parent_group ID로 필터링
        artist_id: int | None = None,
        category: EventCategory | None = None,
        visibility: EventVisibility | None = None,
        is_active: bool = True,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
    ) -> tuple[list[Events], int]:
        """이벤트 목록 조회"""
        query = Events.filter(is_active=is_active)

        # 🔹 artist_parent_group 필터 (Artist의 parent_group을 통해 필터링)
        if artist_parent_group is not None:
            query = query.filter(artist__parent_group=artist_parent_group)
        if artist_id:
            query = query.filter(artist_id=artist_id)
        if category:
            query = query.filter(category=category)
        if visibility:
            query = query.filter(visibility=visibility)
        if start_date:
            query = query.filter(start_time__gte=start_date)
        if end_date:
            query = query.filter(start_time__lte=end_date)

        total = await query.count()
        events = (
            await query.select_related("artist")  # 🔹 Artist만 join
            .offset(skip)
            .limit(limit)
            .order_by("-start_time")
        )

        return events, total

    @staticmethod
    async def get_by_id(event_id: int) -> Events | None:
        """이벤트 상세 조회"""
        try:
            return await Events.get(id=event_id).select_related("artist")
        except DoesNotExist:
            return None

    @staticmethod
    async def bulk_create(events_data: list[dict]) -> tuple[int, list[str]]:
        """일괄 이벤트 생성 - 트랜잭션 처리"""
        from tortoise.transactions import in_transaction

        created_count = 0
        errors = []

        try:
            async with in_transaction() as connection:
                for i, event_data in enumerate(events_data):
                    try:
                        # Artist 존재 확인
                        await Artist.get(
                            id=event_data["artist_id"], using_db=connection
                        )

                        await Events.create(**event_data, using_db=connection)
                        created_count += 1
                    except DoesNotExist:
                        errors.append(f"Row {i + 1}: Artist not found")
                    except Exception as e:
                        errors.append(f"Row {i + 1}: {str(e)}")
        except Exception as e:
            errors.append(f"Transaction failed: {str(e)}")

        return created_count, errors

    @staticmethod
    async def get_upcoming_events(hours_ahead: int = 1) -> list[Events]:
        """예정된 이벤트 조회 (알림용)"""
        now = datetime.now()
        target_time = now + timedelta(hours=hours_ahead)

        if hours_ahead == 1:
            return await Events.filter(
                start_time__lte=target_time,
                start_time__gt=now,
                one_hour_notification_sent=False,
                is_active=True,
                visibility__in=[
                    EventVisibility.PUBLIC,
                    EventVisibility.SUBSCRIBERS_ONLY,
                ],
            ).select_related("artist")
        else:
            return await Events.filter(
                instant_notification_sent=False,
                is_active=True,
                visibility__in=[
                    EventVisibility.PUBLIC,
                    EventVisibility.SUBSCRIBERS_ONLY,
                ],
            ).select_related("artist")

    @staticmethod
    async def get_events_by_date_range(
        start_date: datetime, end_date: datetime
    ) -> list[Events]:
        """날짜 범위로 이벤트 조회"""
        return (
            await Events.filter(
                start_time__gte=start_date, start_time__lte=end_date, is_active=True
            )
            .select_related("artist")
            .order_by("start_time")
        )

    @staticmethod
    async def mark_notification_sent(event_id: int, notification_type: str) -> bool:
        """알림 발송 상태 업데이트"""
        try:
            event = await Events.get(id=event_id)
            if notification_type == "instant":
                event.instant_notification_sent = True
            elif notification_type == "one_hour":
                event.one_hour_notification_sent = True
            await event.save()
            return True
        except DoesNotExist:
            return False

# author : Juwon
# date : 2025.09.22
# content : 기존의 services.py 부분

class EventService:
    """이벤트 비즈니스 로직 서비스"""

    @staticmethod
    async def process_upload_file(file: UploadFile) -> FileUploadResponse:
        """파일 업로드 처리"""
        try:
            contents = await file.read()

            if file.filename.endswith(".xlsx"):
                df = pd.read_excel(io.BytesIO(contents))
            elif file.filename.endswith(".csv"):
                df = pd.read_csv(io.StringIO(contents.decode("utf-8")))
            else:
                raise ValueError("Unsupported file format")

            # 필수 컬럼 검증
            required_columns = ["artist_id", "title", "start_time", "category"]
            missing_columns = set(required_columns) - set(df.columns)
            if missing_columns:
                raise ValueError(f"Missing columns: {', '.join(missing_columns)}")

            # 데이터 변환 및 검증
            events_data = []
            errors = []

            for index, row in df.iterrows():
                try:
                    start_time = pd.to_datetime(row["start_time"])
                    end_time = (
                        pd.to_datetime(row["end_time"])
                        if pd.notna(row.get("end_time"))
                        else None
                    )

                    event_data = {
                        "artist_id": int(row["artist_id"]),
                        "title": str(row["title"]).strip(),
                        "description": str(row.get("description", "")) or None,
                        "start_time": start_time,
                        "end_time": end_time,
                        "location": str(row.get("location", "")) or None,
                        "category": EventCategory(row["category"]),
                        "visibility": EventVisibility(
                            row.get("visibility", EventVisibility.PUBLIC)
                        ),
                    }

                    # 데이터 검증
                    if len(event_data["title"]) > 200:
                        raise ValueError("Title too long (max 200 characters)")
                    if event_data["location"] and len(event_data["location"]) > 200:
                        raise ValueError("Location too long (max 200 characters)")
                    if end_time and start_time >= end_time:
                        raise ValueError("End time must be after start time")

                    events_data.append(event_data)
                except Exception as e:
                    errors.append(f"Row {index + 2}: {str(e)}")

            # 일괄 생성
            created_count, creation_errors = await EventCRUD.bulk_create(events_data)
            all_errors = errors + creation_errors

            logger.info(
                f"File upload processed: {created_count}/{len(df)} events created"
            )

            return FileUploadResponse(
                message="File processed successfully",
                total_processed=len(df),
                successful=created_count,
                failed=len(df) - created_count,
                errors=all_errors[:10],
            )

        except Exception as e:
            raise HTTPException(
                status_code=400, detail=f"File processing error: {str(e)}"
            ) from e

    @staticmethod
    async def generate_template() -> BinaryIO:
        """업로드 템플릿 생성"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Events Template"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        headers = [
            "artist_id",
            "title",
            "description",
            "start_time",
            "end_time",
            "location",
            "category",
            "visibility",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 설명 시트 추가
        ws_info = wb.create_sheet("Information")
        info_data = [
            ["Column", "Description", "Required", "Example"],
            ["artist_id", "아티스트 ID (정수)", "Yes", "1"],
            ["title", "이벤트 제목 (최대 200자)", "Yes", "2024 콘서트"],
            ["description", "이벤트 설명", "No", "특별한 콘서트 이벤트"],
            [
                "start_time",
                "시작 시간 (YYYY-MM-DD HH:MM:SS)",
                "Yes",
                "2024-12-25 19:00:00",
            ],
            [
                "end_time",
                "종료 시간 (YYYY-MM-DD HH:MM:SS)",
                "No",
                "2024-12-25 22:00:00",
            ],
            ["location", "위치 (최대 200자)", "No", "올림픽공원"],
            ["category", "카테고리", "Yes", "concert, fanmeeting, showcase, etc."],
            ["visibility", "공개 설정", "No", "public, private, subscribers_only"],
        ]

        for row, data in enumerate(info_data, 1):
            for col, value in enumerate(data, 1):
                cell = ws_info.cell(row=row, column=col, value=value)
                if row == 1:  # 헤더
                    cell.font = header_font
                    cell.fill = header_fill

        # 샘플 데이터 (메인 시트에)
        sample_data = [
            [
                1,
                "Sample Concert",
                "Sample Description",
                "2024-12-25 19:00:00",
                "2024-12-25 22:00:00",
                "Seoul Olympic Park",
                "concert",
                "public",
            ],
            [
                1,
                "Fan Meeting",
                "Meet & Greet Event",
                "2024-12-30 15:00:00",
                "2024-12-30 17:00:00",
                "Coex Hall",
                "fanmeeting",
                "subscribers_only",
            ],
        ]

        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)

        # 컬럼 너비 조정
        for sheet in [ws, ws_info]:
            for column in sheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column)
                sheet.column_dimensions[column[0].column_letter].width = min(
                    max_length + 2, 25
                )

        # 바이트 스트림으로 변환
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        return stream

    @staticmethod
    async def export_to_excel(
        artist_id: int | None = None,
        category: EventCategory | None = None,
        start_date: str | None = None,
        end_date: str | None = None,
    ) -> BinaryIO:
        """이벤트 데이터 엑셀 내보내기"""
        # 날짜 파싱
        start_dt = datetime.fromisoformat(start_date) if start_date else None
        end_dt = datetime.fromisoformat(end_date) if end_date else None

        # 데이터 조회
        events, _ = await EventCRUD.get_list(
            limit=10000,
            artist_id=artist_id,
            category=category,
            start_date=start_dt,
            end_date=end_dt,
        )

        # 데이터 변환
        data = []
        for event in events:
            data.append(
                {
                    "ID": event.id,
                    "Artist ID": event.artist_id,
                    "Title": event.title,
                    "Description": event.description or "",
                    "Start Time": event.start_time.strftime("%Y-%m-%d %H:%M:%S"),
                    "End Time": event.end_time.strftime("%Y-%m-%d %H:%M:%S")
                    if event.end_time
                    else "",
                    "Location": event.location or "",
                    "Category": event.category.value,
                    "Visibility": event.visibility.value,
                    "Instant Notification Sent": event.instant_notification_sent,
                    "One Hour Notification Sent": event.one_hour_notification_sent,
                    "Is Active": event.is_active,
                    "Created At": event.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                    "Updated At": event.updated_at.strftime("%Y-%m-%d %H:%M:%S"),
                }
            )

        # DataFrame 생성 및 엑셀 변환
        df = pd.DataFrame(data)

        stream = io.BytesIO()
        with pd.ExcelWriter(stream, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Events", index=False)

            # 스타일링
            _ = writer.book  # workbook variable not used
            worksheet = writer.sheets["Events"]

            # 헤더 스타일
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 컬럼 너비 조정
            for column in worksheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column)
                worksheet.column_dimensions[column[0].column_letter].width = min(
                    max_length + 2, 25
                )

        stream.seek(0)
        logger.info(f"Excel export completed: {len(data)} events exported")
        return stream

    @staticmethod
    async def trigger_notifications():
        """알림 트리거 - 더 안전한 처리"""
        try:
            # 데이터베이스 연결 확인
            from tortoise import Tortoise

            if not Tortoise.get_connection("default"):
                logger.warning(
                    "Database connection not available, skipping notifications"
                )
                return

            # 즉시 알림
            instant_events = await EventCRUD.get_upcoming_events(hours_ahead=0)
            instant_success = 0

            for event in instant_events:
                try:
                    await notification_service.send_instant_notification(event)
                    await EventCRUD.mark_notification_sent(event.id, "instant")
                    instant_success += 1
                except Exception as e:
                    logger.error(
                        f"Failed to send instant notification for event {event.id}: {str(e)}"
                    )

            # 1시간 전 알림
            upcoming_events = await EventCRUD.get_upcoming_events(hours_ahead=1)
            upcoming_success = 0

            for event in upcoming_events:
                try:
                    await notification_service.send_upcoming_notification(event)
                    await EventCRUD.mark_notification_sent(event.id, "one_hour")
                    upcoming_success += 1
                except Exception as e:
                    logger.error(
                        f"Failed to send upcoming notification for event {event.id}: {str(e)}"
                    )

            logger.info(
                f"Notifications completed: {instant_success}/{len(instant_events)} instant, {upcoming_success}/{len(upcoming_events)} upcoming"
            )

        except Exception as e:
            logger.error(f"Notification trigger failed: {str(e)}")

    @staticmethod
    async def get_event_statistics():
        """이벤트 통계 조회"""
        # 전체 이벤트 수
        total_events, _ = await EventCRUD.get_list(limit=1)

        # 카테고리별 통계 (추후 구현 예정)
        return {"total_events": len(total_events)}

# author : Juwon
# date : 2025.09.22
# content : notification.py를 services.py로 이동




class NotificationService:
    """알림 서비스 클래스"""

    def __init__(self):
        # 설정값들 (실제 환경에서는 config에서 가져오기)
        self.push_service_url = "http://localhost:8001"  # 푸시 서비스 URL
        self.email_service_url = "http://localhost:8002"  # 이메일 서비스 URL
        self.webhook_urls = [
            # "https://discord.com/api/webhooks/your-webhook-url",
            # "https://hooks.slack.com/services/your-webhook-url"
        ]

    async def send_instant_notification(self, event: Events):
        """즉시 알림 전송 (이벤트 등록시)"""
        message = self._create_instant_message(event)

        # 여러 채널로 동시 전송
        tasks = [
            self._send_push_notification(message),
            self._send_email_notification(message),
            self._send_webhook_notifications(message, event),
        ]

        results = await asyncio.gather(*tasks, return_exceptions=True)

        # 로깅
        success_count = sum(
            1 for result in results if not isinstance(result, Exception)
        )
        logger.info(
            f"Instant notification sent for event {event.id}: {success_count}/{len(tasks)} channels succeeded"
        )

    async def send_upcoming_notification(self, event: Events):
        """1시간 전 알림 전송"""
        message = self._create_upcoming_message(event)

        tasks = [
            self._send_push_notification(message),
            self._send_email_notification(message),
        ]

        results = await asyncio.gather(*tasks, return_exceptions=True)

        success_count = sum(
            1 for result in results if not isinstance(result, Exception)
        )
        logger.info(
            f"Upcoming notification sent for event {event.id}: {success_count}/{len(tasks)} channels succeeded"
        )

    def _create_instant_message(self, event: Events) -> dict[str, Any]:
        """즉시 알림 메시지 생성 - 안전성 개선"""
        artist_name = "Unknown Artist"  # 기본값

        if hasattr(event, "artist") and event.artist:
            artist_name = getattr(event.artist, "name", "Unknown Artist")

        return {
            "type": "instant",
            "event_id": event.id,
            "title": f"🎵 새로운 이벤트: {event.title}",
            "body": f"{artist_name}의 {event.category.value} 일정이 등록되었습니다!",
            "data": {
                "event_id": event.id,
                "artist_id": event.artist_id,
                "start_time": event.start_time.isoformat(),
                "location": event.location or "미정",
                "category": event.category.value,
            },
        }

    def _create_upcoming_message(self, event: Events) -> dict[str, Any]:
        """1시간 전 알림 메시지 생성 - 안전성 개선"""
        artist_name = "Unknown Artist"  # 기본값

        if hasattr(event, "artist") and event.artist:
            artist_name = getattr(event.artist, "name", "Unknown Artist")

        return {
            "type": "upcoming",
            "event_id": event.id,
            "title": f"⏰ 곧 시작: {event.title}",
            "body": f"{artist_name}의 {event.category.value}이 1시간 후 시작됩니다!",
            "data": {
                "event_id": event.id,
                "artist_id": event.artist_id,
                "start_time": event.start_time.isoformat(),
                "location": event.location or "미정",
                "category": event.category.value,
            },
        }

    async def _send_push_notification(self, message: dict[str, Any]) -> bool:
        """푸시 알림 전송"""
        if not self.push_service_url:
            logger.warning("Push service URL not configured")
            return False

        try:
            async with httpx.AsyncClient() as client:
                response = await client.post(
                    f"{self.push_service_url}/send", json=message, timeout=10.0
                )
                response.raise_for_status()
                logger.info(
                    f"Push notification sent successfully for event {message['event_id']}"
                )
                return True

        except httpx.RequestError as e:
            logger.error(f"Push notification request failed: {str(e)}")
            return False
        except httpx.HTTPStatusError as e:
            logger.error(f"Push notification HTTP error: {e.response.status_code}")
            return False
        except Exception as e:
            logger.error(f"Failed to send push notification: {str(e)}")
            return False

    async def _send_email_notification(self, message: dict[str, Any]) -> bool:
        """이메일 알림 전송"""
        if not self.email_service_url:
            logger.warning("Email service URL not configured")
            return False

        try:
            email_data = {
                "subject": message["title"],
                "body": message["body"],
                "template": "event_notification",
                "data": message["data"],
            }

            async with httpx.AsyncClient() as client:
                response = await client.post(
                    f"{self.email_service_url}/send", json=email_data, timeout=10.0
                )
                response.raise_for_status()
                logger.info(
                    f"Email notification sent successfully for event {message['event_id']}"
                )
                return True

        except httpx.RequestError as e:
            logger.error(f"Email notification request failed: {str(e)}")
            return False
        except httpx.HTTPStatusError as e:
            logger.error(f"Email notification HTTP error: {e.response.status_code}")
            return False
        except Exception as e:
            logger.error(f"Failed to send email notification: {str(e)}")
            return False

    async def _send_webhook_notifications(
        self, message: dict[str, Any], event: Events
    ) -> bool:
        """웹훅 알림 전송 - artist 안전성 체크 추가"""
        if not self.webhook_urls:
            logger.info("No webhook URLs configured")
            return True  # 설정이 없어도 에러는 아님

        # artist 이름 안전하게 가져오기
        artist_name = "Unknown Artist"
        if hasattr(event, "artist") and event.artist:
            artist_name = getattr(event.artist, "name", "Unknown Artist")

        # 디스코드용 웹훅 페이로드
        webhook_payload = {
            "content": f"🎵 **{message['title']}**\n{message['body']}",
            "embeds": [
                {
                    "title": event.title,
                    "description": event.description or "설명이 없습니다.",
                    "color": 0x667EEA,
                    "fields": [
                        {"name": "아티스트", "value": artist_name, "inline": True},
                        {
                            "name": "카테고리",
                            "value": event.category.value,
                            "inline": True,
                        },
                        {
                            "name": "시작 시간",
                            "value": event.start_time.strftime("%Y-%m-%d %H:%M"),
                            "inline": True,
                        },
                        {
                            "name": "위치",
                            "value": event.location or "미정",
                            "inline": True,
                        },
                    ],
                    "timestamp": datetime.utcnow().isoformat(),
                }
            ],
        }

        successful = 0
        for webhook_url in self.webhook_urls:
            try:
                async with httpx.AsyncClient() as client:
                    response = await client.post(
                        webhook_url, json=webhook_payload, timeout=10.0
                    )
                    response.raise_for_status()
                    successful += 1

            except Exception as e:
                logger.error(
                    f"Failed to send webhook notification to {webhook_url}: {str(e)}"
                )

        logger.info(
            f"Webhook notifications sent: {successful}/{len(self.webhook_urls)} successful"
        )
        return successful > 0

    async def send_batch_notification(
        self, events: list[Events], notification_type: str = "batch"
    ):
        """일괄 알림 전송 (관리자용)"""
        if not events:
            logger.info("No events to send batch notification for")
            return

        # 안전하게 아티스트 이름 가져오기
        event_data = []
        for event in events[:5]:  # 최대 5개만
            artist_name = "Unknown Artist"
            if hasattr(event, "artist") and event.artist:
                artist_name = getattr(event.artist, "name", "Unknown Artist")

            event_data.append(
                {
                    "id": event.id,
                    "title": event.title,
                    "start_time": event.start_time.isoformat(),
                    "artist_name": artist_name,
                }
            )

        message = {
            "type": notification_type,
            "title": f"📅 {len(events)}개의 이벤트 알림",
            "body": f"새로운 {len(events)}개의 이벤트가 추가되었습니다.",
            "data": {"event_count": len(events), "events": event_data},
        }

        await self._send_push_notification(message)
        logger.info(f"Batch notification sent for {len(events)} events")

    async def send_test_notification(self, message: str = "테스트 알림입니다."):
        """테스트 알림 전송 (개발/디버깅용)"""
        test_message = {
            "type": "test",
            "event_id": 0,
            "title": "🧪 테스트 알림",
            "body": message,
            "data": {"timestamp": datetime.now().isoformat()},
        }

        tasks = [
            self._send_push_notification(test_message),
            self._send_email_notification(test_message),
        ]

        results = await asyncio.gather(*tasks, return_exceptions=True)
        success_count = sum(
            1 for result in results if not isinstance(result, Exception)
        )

        logger.info(
            f"Test notification sent: {success_count}/{len(tasks)} channels succeeded"
        )
        return success_count > 0


# 싱글톤 인스턴스 생성
notification_service = NotificationService()



# author : Juwon
# date : 2025.09.25
# content : 크롤링 관련 메서드들 추가

@staticmethod
async def import_scraped_events(scraped_events: list[dict], source: str) -> FileUploadResponse:
    """크롤링된 이벤트를 DB에 저장"""
    try:
        events_data = []
        errors = []

        for i, event in enumerate(scraped_events):
            try:
                # 날짜 파싱
                start_time = EventService._parse_scraped_date(
                    event.get('date'),
                    event.get('time')
                )

                # 카테고리 매핑
                category = EventService._map_scraped_category(event.get('type', ''))

                # 아티스트 ID 찾기 또는 생성
                artist_id = await EventService._get_or_create_artist_for_scraped_event(
                    event.get('artist'),
                    source
                )

                if not artist_id:
                    errors.append(f"Row {i + 1}: Artist not found or created")
                    continue

                event_data = {
                    "artist_id": artist_id,
                    "title": event.get('title', '').strip()[:200],  # 길이 제한
                    "description": f"Source: {source}\n" + (event.get('description') or ''),
                    "start_time": start_time,
                    "end_time": None,  # 크롤링 데이터에는 종료시간이 없는 경우가 많음
                    "location": (event.get('location') or '')[:200] if event.get('location') else None,
                    "category": category,
                    "visibility": EventVisibility.PUBLIC,
                }

                # 중복 체크
                if not await EventService._is_duplicate_event(
                        artist_id, event_data['title'], start_time
                ):
                    events_data.append(event_data)
                else:
                    logger.info(f"Duplicate event skipped: {event_data['title']}")

            except Exception as e:
                errors.append(f"Row {i + 1}: {str(e)}")

        # 일괄 생성
        created_count, creation_errors = await EventCRUD.bulk_create(events_data)
        all_errors = errors + creation_errors

        logger.info(f"Scraped events import: {created_count}/{len(scraped_events)} created from {source}")

        return FileUploadResponse(
            message=f"Scraped events from {source} processed",
            total_processed=len(scraped_events),
            successful=created_count,
            failed=len(scraped_events) - created_count,
            errors=all_errors[:10]
        )

    except Exception as e:
        logger.error(f"Scraped events import error: {str(e)}")
        raise HTTPException(
            status_code=400,
            detail=f"Import processing error: {str(e)}"
        )


@staticmethod
def _parse_scraped_date(date_str: str, time_str: str = None) -> datetime:
    """크롤링된 날짜 문자열을 datetime으로 변환"""
    if not date_str or date_str == "날짜 없음" or date_str == "날짜 미상":
        return datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    try:
        # 다양한 날짜 형식 처리
        if re.match(r'\d{4}-\d{2}-\d{2}', date_str):
            base_date = datetime.strptime(date_str[:10], '%Y-%m-%d')
        elif re.match(r'\d{1,2}월\s*\d{1,2}일', date_str):
            # "12월 25일" 형태
            month_day = re.search(r'(\d{1,2})월\s*(\d{1,2})일', date_str)
            if month_day:
                month, day = month_day.groups()
                current_year = datetime.now().year
                base_date = datetime(current_year, int(month), int(day))
            else:
                base_date = datetime.now()
        else:
            base_date = datetime.now()

        # 시간 정보 추가
        if time_str:
            time_match = re.search(r'(\d{1,2}):(\d{2})', time_str)
            if time_match:
                hour, minute = time_match.groups()
                base_date = base_date.replace(hour=int(hour), minute=int(minute))

        return base_date

    except Exception as e:
        logger.warning(f"Date parsing failed for '{date_str}': {str(e)}")
        return datetime.now()


@staticmethod
def _map_scraped_category(type_str: str) -> EventCategory:
    """크롤링된 타입을 EventCategory로 매핑"""
    if not type_str:
        return EventCategory.OTHER

    type_lower = type_str.lower()

    mapping = {
        'concert': EventCategory.CONCERT,
        '콘서트': EventCategory.CONCERT,
        'fanmeeting': EventCategory.FANMEETING,
        '팬미팅': EventCategory.FANMEETING,
        'showcase': EventCategory.SHOWCASE,
        '쇼케이스': EventCategory.SHOWCASE,
        'broadcast': EventCategory.BROADCAST,
        '방송': EventCategory.BROADCAST,
        'variety': EventCategory.VARIETY,
        '예능': EventCategory.VARIETY,
        'music': EventCategory.MUSIC,
        '음악': EventCategory.MUSIC,
        'release': EventCategory.RELEASE,
        '발매': EventCategory.RELEASE,
        'comeback': EventCategory.COMEBACK,
        '컴백': EventCategory.COMEBACK
    }

    for key, value in mapping.items():
        if key in type_lower:
            return value

    return EventCategory.OTHER


@staticmethod
async def _get_or_create_artist_for_scraped_event(artist_name: str, source: str) -> int | None:
    """크롤링된 이벤트의 아티스트 ID 찾기 또는 생성"""
    if not artist_name:
        return 1  # 기본 아티스트 ID

    try:
        from app.features.artists.models import Artist

        artist = await Artist.filter(name__icontains=artist_name).first()
        if artist:
            return artist.id

        # 기본 아티스트 ID 반환
        logger.warning(f"Artist '{artist_name}' not found, using default artist ID")
        return 1

    except Exception as e:
        logger.error(f"Error getting artist '{artist_name}': {str(e)}")
        return 1


@staticmethod
async def _is_duplicate_event(artist_id: int, title: str, start_time: datetime) -> bool:
    """중복 이벤트 체크"""
    try:
        from app.features.events.models import Events

        # 같은 아티스트, 제목, 날짜가 같은 이벤트가 있는지 확인
        start_date = start_time.date()
        end_date = datetime.combine(start_date, datetime.max.time())

        existing = await Events.filter(
            artist_id=artist_id,
            title=title,
            start_time__gte=start_date,
            start_time__lte=end_date,
            is_active=True
        ).first()

        return existing is not None

    except Exception as e:
        logger.error(f"Error checking duplicate event: {str(e)}")
        return False